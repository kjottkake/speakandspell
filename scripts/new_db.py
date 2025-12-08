import pandas as pd
from pymongo import MongoClient
from dotenv import load_dotenv
import os
from openpyxl import load_workbook

load_dotenv()

DB_URL = os.getenv("DB_URL")
client = MongoClient(DB_URL)
db = client["speaknspell"]

trial_collection = db["trials"]
bin_collection = db["bins"]
exercise_collection = db["exercises"]
word_collection = db["words"]
rule_collection = db["rules"]

trial_collection.drop()
bin_collection.drop()
exercise_collection.drop()
word_collection.drop()
rule_collection.drop()

bin_id = 0
bins = []

exercise_id = 0
exercises = []

trial_id = 0
trials = []

words = []
rules = []

def read_rule(rule: str):
    period = rule.find(".")
    if period == -1:
        return rule
    start = period + 1
    while start < len(rule) and rule[start] == " ":
        start += 1
    return rule[start:]

words_df = pd.read_excel("Localisation word material_example.xlsx", sheet_name="Sheet1", header=1)
for column in words_df.columns:
    if "Unnamed" in column:
        words_df.drop(column, axis=1, inplace=True)
words_df.dropna(how="all", inplace=True)

word_wb = load_workbook("Localisation word material_example.xlsx")
word_ws = word_wb.active
for i, (_, row) in enumerate(words_df.iterrows()):
    word = {"_id": i}
    for l in [
        ["N", "norwegian", 1],
        ["S", "swedish", 5],
        ["D", "danish", 9]
    ]:
        cell = word_ws.cell(row=i+4, column=l[2])
        word[l[1]] = {
            "word": row[f"Words_{l[0]}"],
            "pronunciation": row[f"Pronunc_{l[0]}"],
            "sound": row[f"Sound_{l[0]}"],
            "english": row[f"Words_{l[0]}-E"],
            "is_cognate": cell.fill.fgColor.rgb.upper in ["FF0000FF", "0000FF"]
        }
    words.append(word)

rules_df = pd.read_excel("Localisation rules_example.xlsx", sheet_name="Sheet1", header=0)
for column in rules_df.columns:
    if "Unnamed" in column:
        rules_df.drop(column, axis=1, inplace=True)
rules_df.dropna(how="all", inplace=True)

for i, (_, row) in enumerate(rules_df.iterrows()):
    rules.append({
        "_id": i,
        "english": read_rule(row["ENGLISH"]),
        "norwegian": read_rule(row["NORSK"]),
        "swedish": read_rule(row["SVENSKA"]),
        "danish": read_rule(row["DANSK"])
    })

for lang in ["norwegian", "swedish", "danish"]:
    for x in [
        ["L2S-C", True, True], 
        ["L2S-V", True, False],
        ["S2L-C", False, True],
        ["S2L-V", False, False]
    ]:
        df = pd.read_excel("Exercises_2025-07-18.xlsx", sheet_name=f"{x[0]} {lang.capitalize()}", header=1)
        for column in df.columns:
            if "Unnamed" in column:
                df.drop(column, axis=1, inplace=True)
        if "Comments"in df:
            df.drop("Comments", axis=1, inplace=True)
        df.dropna(how="all", inplace=True)

        current_exercise = 0
        current_bin = 0
        for _, row in df.iterrows():
            word = next((i for i, word in enumerate(words) if word[lang]["word"] == row["Words"]), None)
            rule = next((i for i, rule in enumerate(rules) if rule["english"] == row["Rule"]), None)
            if word is None or rule is None:
                continue
            if row["Exercise"] != current_exercise:
                current_exercise = row["Exercise"]
                exercise_id += 1
                exercises.append({
                    "_id": exercise_id,
                    "target": row["Target"],
                    "language": lang,
                    "is_speak": x[1],
                    "is_consonant": x[2]
                })
            if row["Bin"] != current_bin:
                current_bin = row["Bin"]
                bin_id += 1
                bins.append({
                    "_id": bin_id,
                    "rule": rule,
                    "context": row["Context"],
                    "is_exception": "-x" in str(current_bin),
                    "exercise": exercise_id
                })
            trial_id += 1
            trials.append({
                "_id": trial_id,
                "bin": bin_id,
                "word": word
            })

exercise_collection.insert_many(exercises)
bin_collection.insert_many(bins)
trial_collection.insert_many(trials)
word_collection.insert_many(words)
rule_collection.insert_many(rules)